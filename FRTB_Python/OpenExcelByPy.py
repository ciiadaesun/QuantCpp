#%%
"""
v1.0.2
"""
import pandas as pd
import win32com.client as win32
from packaging import version
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import datetime
curdir = os.getcwd()

print("###########################################")
print("### 사용중인 엑셀을 종료해 주시기 바랍니다. #####")
print("###########################################")

def remove_timezone_from_df(df):
    for col in df.columns:
        df[col] = df[col].map(
            lambda x: x.replace(tzinfo=None)
            if hasattr(x, "tzinfo") and x.tzinfo is not None
            else x
        )
    return df

def excel_horizontal_to_openpyxl(value):

    mapping = {
        -4131: "left",
        -4108: "center",
        -4152: "right",
        -4130: "justify"
    }

    return mapping.get(value, None)


def excel_vertical_to_openpyxl(value):

    mapping = {
        -4160: "top",
        -4108: "center",
        -4107: "bottom"
    }

    return mapping.get(value, None)

def normalize_number_format(fmt, cell_type):
    fmt = str(fmt) if fmt is not None else "General"

    # 텍스트
    if cell_type == "text":
        return "@"

    # 날짜
    if cell_type == "date":
        return "yyyy-mm-dd"

    # 숫자
    if cell_type == "number":
        # 날짜/시간 포맷이 숫자에 잘못 붙는 것 방지
        lower_fmt = fmt.lower()
        if any(x in lower_fmt for x in ["yy", "mm", "dd", "h:", "am/pm"]):
            return "General"

        # 표준은 그대로 General 처리
        if fmt in ["General", "G/표준", "표준"]:
            return "General"

        # 퍼센트만 유지
        if "%" in fmt:
            return fmt

        # 천단위/소수 포맷은 단순화
        if "," in fmt:
            return "#,##0"

        return "General"

    return "General"

def excel_border_to_openpyxl(line_style, weight):
    if line_style in [-4142, 0, None]:
        return None

    # Excel Weight
    # 1: Hairline, 2: Thin, -4138: Medium, 4: Thick
    if weight == 4:
        return "thick"
    elif weight == -4138:
        return "medium"
    elif weight == 1:
        return "hair"
    elif weight == 2:
        return "thin"

    # fallback
    if line_style == 1:
        return "thin"
    elif line_style == 2:
        return "medium"
    elif line_style == 4:
        return "dashed"
    elif line_style == -4115:
        return "dotted"

    return "thin"

from copy import copy

def apply_border_to_merged_range_from_info(ws, border_infos, r1, c1, r2, c2):

    left_info = border_infos[r1 - 1][c1 - 1]["left"]
    top_info = border_infos[r1 - 1][c1 - 1]["top"]
    right_info = border_infos[r1 - 1][c2 - 1]["right"]
    bottom_info = border_infos[r2 - 1][c1 - 1]["bottom"]

    border = Border(
        left=make_side(left_info),
        top=make_side(top_info),
        right=make_side(right_info),
        bottom=make_side(bottom_info)
    )

    top_left = ws.cell(row=r1, column=c1)
    top_left.border = border

    # 병합셀 Range에 border 전파
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row == r1 and
            merged_range.min_col == c1 and
            merged_range.max_row == r2 and
            merged_range.max_col == c2
        ):
            merged_range.format()
            break

def ole_color_to_rgb(color_value):
    try:
        color_int = int(color_value)

        if color_int < 0:
            return None

        red = color_int % 256
        green = (color_int // 256) % 256
        blue = (color_int // 65536) % 256

        return f"{red:02X}{green:02X}{blue:02X}"
    except:
        return None


def make_side(border_info):
    style = excel_border_to_openpyxl(
        border_info["style"],
        border_info["weight"]
    )
    color = ole_color_to_rgb(border_info["color"])

    if style is None:
        return Side(style=None)

    return Side(style=style, color=color)

#input("다른 엑셀시트를 모두 종료하시기 바랍니다. 종료하고 엔터를 누르세요 : ")
targetfilepath = input("복사하실 엑셀 파일 명을 입력하시오(TempExcel.xlsx) : ")

if '.xlsx' not in targetfilepath:
    targetfilepath += '.xlsx'

if '\\' not in targetfilepath:
    targetfilepath = curdir + '\\' + targetfilepath

resultfilename = input("저장할 파일 명을 입력하시오(ex File2.xlsx) : ")

if ".xlsx" not in resultfilename:
    resultfilename += ".xlsx"

if "\\" not in resultfilename:
    resultfilename = curdir + "\\" + resultfilename

valuestring = input("0: 각각 셀 포멧 그대로복사 or 1: 값만 복사")
valueread = False if len(valuestring) == 0 or valuestring == '0' else True

excel = win32.Dispatch("Excel.Application")
excel.Visible = True
excel.DisplayAlerts = False

wb = excel.Workbooks.Open(targetfilepath)

sheet_name_list = []
data_list = []
merge_info_list = []
col_width_list = []
row_height_list = []
font_info_list = []
alignment_info_list = []
number_format_list = []
value_type_list = []
fill_info_list = []
border_info_list = []
formula_info_list = []

for ws in wb.Worksheets:
    sheet_name_list.append(ws.Name)

for n in range(len(sheet_name_list)):
    print(str(n+1) + "번째 시트 복사 진행합니다.")
    ws = wb.Worksheets(sheet_name_list[n])
    used_range = ws.UsedRange

    start_row = used_range.Row
    start_col = used_range.Column
    nrows = used_range.Rows.Count
    ncols = used_range.Columns.Count

    data = []
    font_infos = []
    alignment_infos = []
    number_format_infos = []
    value_type_infos = []
    fill_infos = []
    border_infos = []    
    formula_infos = []
    for r in range(start_row, start_row + nrows):
        row = []
        font_row = []
        alignment_row = []
        number_format_row = []
        value_type_row = []
        fill_row = []
        border_row = []      
        formula_row = []  
        for c in range(start_col, start_col + ncols):
            cell = ws.Cells(r, c)
            font = cell.Font
            interior = cell.Interior
            font_row.append({
                "name": font.Name,
                "size": font.Size,
                "bold": bool(font.Bold),
                "italic": bool(font.Italic),
                "underline": "single" if font.Underline != -4142 else None,
                "strike": bool(font.Strikethrough),
                "color": font.Color
            })

            alignment_row.append({
                "horizontal": cell.HorizontalAlignment,
                "vertical": cell.VerticalAlignment,
                "wrap": bool(cell.WrapText)
            })

            number_format_row.append(
                cell.NumberFormat
            )

            fill_row.append({
                "color": interior.Color,
                "colorindex": interior.ColorIndex,
                "pattern": interior.Pattern
            })

            border_row.append({
                "left": {
                    "style": cell.Borders(7).LineStyle,
                    "weight": cell.Borders(7).Weight,
                    "color": cell.Borders(7).Color
                },
                "top": {
                    "style": cell.Borders(8).LineStyle,
                    "weight": cell.Borders(8).Weight,
                    "color": cell.Borders(8).Color
                },
                "bottom": {
                    "style": cell.Borders(9).LineStyle,
                    "weight": cell.Borders(9).Weight,
                    "color": cell.Borders(9).Color
                },
                "right": {
                    "style": cell.Borders(10).LineStyle,
                    "weight": cell.Borders(10).Weight,
                    "color": cell.Borders(10).Color
                }
            })

            formula_row.append(bool(cell.HasFormula))

            # 병합셀 내부 셀은 pandas 저장 시 값 중복 방지
            if cell.MergeCells:
                ma = cell.MergeArea
                if cell.Row != ma.Row or cell.Column != ma.Column:
                    value = None
                else:
                    if valueread:
                        value = cell.Value
                    else:
                        value = cell.Formula if cell.HasFormula else cell.Value
                        if isinstance(value, (datetime.datetime, pd.Timestamp)):
                            if value.hour == 0 and value.minute == 0 and value.second == 0:
                                value = value.date()

            else:
                if valueread:
                    value = cell.Value
                else:
                    value = cell.Formula if cell.HasFormula else cell.Value
                    if isinstance(value, (datetime.datetime, pd.Timestamp)):
                        if value.hour == 0 and value.minute == 0 and value.second == 0:
                            value = value.date()

            if isinstance(value, str):
                value_type_row.append("text")
            elif isinstance(value, (int, float)):
                value_type_row.append("number")
            elif isinstance(value, (datetime.datetime, datetime.date, pd.Timestamp)):
                value_type_row.append("date")
            else:
                value_type_row.append("other")

            row.append(value)
        font_infos.append(font_row)
        alignment_infos.append(alignment_row)
        number_format_infos.append(number_format_row)
        value_type_infos.append(value_type_row)
        fill_infos.append(fill_row)
        border_infos.append(border_row)
        formula_infos.append(formula_row)

        data.append(row)
    font_info_list.append(font_infos)
    alignment_info_list.append(alignment_infos)
    number_format_list.append(number_format_infos)
    value_type_list.append(value_type_infos)
    fill_info_list.append(fill_infos)
    border_info_list.append(border_infos) 
    formula_info_list.append(formula_infos)

    df = pd.DataFrame(data)
    df = remove_timezone_from_df(df)
    data_list.append(df)

    # 병합셀 정보 저장
    merge_ranges = []
    merge_address_set = set()

    for r in range(start_row, start_row + nrows):
        for c in range(start_col, start_col + ncols):
            cell = ws.Cells(r, c)

            if cell.MergeCells:
                ma = cell.MergeArea
                addr = ma.Address

                # 같은 병합영역 중복 방지
                if addr not in merge_address_set:
                    merge_address_set.add(addr)

                    r1 = ma.Row - start_row + 1
                    c1 = ma.Column - start_col + 1
                    r2 = r1 + ma.Rows.Count - 1
                    c2 = c1 + ma.Columns.Count - 1

                    merge_ranges.append((r1, c1, r2, c2))

    merge_info_list.append(merge_ranges)

    # 열 너비 저장
    col_widths = []
    for c in range(start_col, start_col + ncols):
        col_widths.append(ws.Columns(c).ColumnWidth)

    col_width_list.append(col_widths)

    # 행 높이 저장
    row_heights = []
    for r in range(start_row, start_row + nrows):
        row_heights.append(ws.Rows(r).RowHeight)

    row_height_list.append(row_heights)

    print(
        sheet_name_list[n] + " 복사완료, 잠시만기다려주세요"
        if n < len(sheet_name_list) - 1
        else "복사완료"
    )
print("50%완료")
# 1차 저장
writer = pd.ExcelWriter(resultfilename, engine="openpyxl")

for n in range(len(sheet_name_list)):
    data_list[n].to_excel(
        writer,
        sheet_name=sheet_name_list[n],
        index=False,
        header=False
    )

writer.close()

# 2차로 병합셀 / 열너비 / 행높이 적용
out_wb = load_workbook(resultfilename)

for n, sheet_name in enumerate(sheet_name_list):
    out_ws = out_wb[sheet_name]

    # 폰트 / 글자색 적용
    for r_idx, font_row in enumerate(font_info_list[n], start=1):
        for c_idx, font_info in enumerate(font_row, start=1):

            out_cell = out_ws.cell(row=r_idx, column=c_idx)

            # 병합셀 내부는 쓰기 불가인 경우가 있어서 skip
            if out_cell.__class__.__name__ == "MergedCell":
                continue

            color = None

            if font_info["color"] is not None:
                try:
                    # Excel COM 색상값은 BGR 정수라서 RGB로 변환
                    color_int = int(font_info["color"])
                    red = color_int % 256
                    green = (color_int // 256) % 256
                    blue = (color_int // 65536) % 256
                    color = f"{red:02X}{green:02X}{blue:02X}"
                except:
                    color = None

            out_cell.font = Font(
                name=font_info["name"],
                size=font_info["size"],
                bold=font_info["bold"],
                italic=font_info["italic"],
                underline=font_info["underline"],
                strike=font_info["strike"],
                color=color
            )    

    # 배경색 적용
    for r_idx, fill_row in enumerate(fill_info_list[n], start=1):
        for c_idx, fill_info in enumerate(fill_row, start=1):

            out_cell = out_ws.cell(row=r_idx, column=c_idx)

            if out_cell.__class__.__name__ == "MergedCell":
                continue

            try:

                # 색상 없음
                if fill_info["colorindex"] in [-4142, -4105]:
                    continue

                if fill_info["pattern"] in [-4142]:
                    continue

                color_int = int(fill_info["color"])

                red = color_int % 256
                green = (color_int // 256) % 256
                blue = (color_int // 65536) % 256

                rgb = f"{red:02X}{green:02X}{blue:02X}"

                out_cell.fill = PatternFill(
                    fill_type="solid",
                    start_color=rgb,
                    end_color=rgb
                )

            except:
                pass

    # 정렬 적용
    for r_idx, align_row in enumerate(alignment_info_list[n], start=1):
        for c_idx, align_info in enumerate(align_row, start=1):

            out_cell = out_ws.cell(row=r_idx, column=c_idx)

            if out_cell.__class__.__name__ == "MergedCell":
                continue

            out_cell.alignment = Alignment(
                horizontal=excel_horizontal_to_openpyxl(
                    align_info["horizontal"]
                ),
                vertical=excel_vertical_to_openpyxl(
                    align_info["vertical"]
                ),
                wrap_text=align_info["wrap"]
            )

    # 표시형식 적용
    for r_idx, format_row in enumerate(number_format_list[n], start=1):
        for c_idx, fmt in enumerate(format_row, start=1):

            out_cell = out_ws.cell(row=r_idx, column=c_idx)
            is_formula = formula_info_list[n][r_idx - 1][c_idx - 1]
            if out_cell.__class__.__name__ == "MergedCell":
                continue

            try:
                if is_formula :     
                    out_cell.number_format = str(fmt) if fmt is not None else "General"
                else : 
                    cell_type = value_type_list[n][r_idx - 1][c_idx - 1]
                    out_cell.number_format = normalize_number_format(fmt, cell_type)
            except:
                out_cell.number_format = "General"

    # 테두리 적용
    for r_idx, border_row in enumerate(border_info_list[n], start=1):
        for c_idx, border_info in enumerate(border_row, start=1):

            out_cell = out_ws.cell(row=r_idx, column=c_idx)

            if out_cell.__class__.__name__ == "MergedCell":
                continue

            out_cell.border = Border(
                left=make_side(border_info["left"]),
                top=make_side(border_info["top"]),
                bottom=make_side(border_info["bottom"]),
                right=make_side(border_info["right"])
            )

        # 병합셀 적용 + 병합셀 외곽 테두리 재적용
        for r1, c1, r2, c2 in merge_info_list[n]:
            out_ws.merge_cells(
                start_row=r1,
                start_column=c1,
                end_row=r2,
                end_column=c2
            )

            apply_border_to_merged_range_from_info(
                out_ws,
                border_info_list[n],
                r1, c1, r2, c2
            )

    # 열 너비 적용
    for idx, width in enumerate(col_width_list[n], start=1):
        col_letter = get_column_letter(idx)
        out_ws.column_dimensions[col_letter].width = width

    # 행 높이 적용
    for idx, height in enumerate(row_height_list[n], start=1):
        out_ws.row_dimensions[idx].height = height

out_wb.save(resultfilename)

wb.Close(SaveChanges=False)
excel.DisplayAlerts = True
excel.Quit()

print("전체 복사 완료")


# %%